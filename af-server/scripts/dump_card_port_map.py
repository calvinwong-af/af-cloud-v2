"""
Temporary helper — dumps city_code -> port_un_code mapping from skipped cards file.
Run from af-server root: .venv\Scripts\python scripts\dump_card_port_map.py
"""
import os
from openpyxl import load_workbook

path = os.path.join(os.path.dirname(__file__), "output", "transport_skipped_cards_20260310.xlsx")
wb = load_workbook(path)
ws = wb.active

# headers: pt_id, port_un_code, city_code, ...
port_map = {}  # city_code -> sorted list of unique port_un_codes
for row in ws.iter_rows(min_row=2, values_only=True):
    city_code = (row[2] or "").strip()
    port_un_code = (row[1] or "").strip()
    if city_code and port_un_code:
        if city_code not in port_map:
            port_map[city_code] = set()
        port_map[city_code].add(port_un_code)

out_path = os.path.join(os.path.dirname(__file__), "output", "card_port_map.txt")
with open(out_path, "w") as f:
    for city_code, ports in sorted(port_map.items()):
        f.write(f"{city_code}|{','.join(sorted(ports))}\n")

print(f"Written {len(port_map)} entries to {out_path}")
