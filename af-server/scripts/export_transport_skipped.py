"""
scripts/export_transport_skipped.py

Export skipped PricingTransport data to Excel for offline reconciliation.

A record is skipped if:
  - city_code is missing (no_city_code)
  - tonnage is missing or not in vehicle_types table (tonnage_not_found)
  - city_code is not found in the areas table (area_not_found)

Trashed records and already-seeded area codes are excluded.

Produces three files in af-server/scripts/output/:
  transport_skipped_areas_YYYYMMDD.xlsx  — unique city_codes with names from Datastore City kind
  transport_skipped_cards_YYYYMMDD.xlsx  — one row per skipped rate card
  transport_skipped_rates_YYYYMMDD.xlsx  — 2024+ rate rows for those skipped cards

The pt_id column is the join key between cards and rates files.

Requires Cloud SQL Auth Proxy running on port 5432.

Run from af-server root with venv active:
    .venv\\Scripts\\python scripts\\export_transport_skipped.py
"""

import logging
import os
import sys
from datetime import date as date_cls

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

from sqlalchemy import text
from core.db import get_engine
from google.cloud import datastore
from google.cloud.datastore.query import PropertyFilter

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

PROJECT_ID = 'cloud-accele-freight'
KIND_PRICING_TRANSPORT = 'PricingTransport'
KIND_MONTHLY_RATE = 'PTMonthlyRateHaulageTransport'
KIND_CITY = 'City'
PT_TRANSPORT = 'PT-TRANSPORT'

_MONTH_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

# Known area codes seeded in migration 026 — treated as resolved, not skipped
SEEDED_AREA_CODES = {"MY-KUL-000", "MY-MLK-000"}

# Legacy port code normalisation (must match migrate_transport_pricing.py)
LEGACY_PORT_TERMINAL_MAP = {
    "MYPKG_N": ("MYPKG", "MYPKG_N"),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_month_year(month_year: str) -> date_cls | None:
    if not month_year or "-" not in month_year:
        return None
    parts = month_year.strip().upper().split("-")
    if len(parts) != 2:
        return None
    month_str, year_str = parts
    month = _MONTH_MAP.get(month_str)
    if not month:
        return None
    try:
        year = int(year_str)
    except ValueError:
        return None
    return date_cls(year, month, 1)


def _safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _get_multi_chunked(ds_client, keys, chunk_size=500) -> list:
    results = []
    for i in range(0, len(keys), chunk_size):
        chunk = keys[i:i + chunk_size]
        entities = ds_client.get_multi(chunk)
        results.extend([e for e in entities if e is not None])
    return results


# ---------------------------------------------------------------------------
# Reference data from DB
# ---------------------------------------------------------------------------

def load_areas_set(conn) -> set[str]:
    rows = conn.execute(text("SELECT area_code FROM areas")).fetchall()
    return {r[0] for r in rows}


def load_known_tonnage_set(conn) -> set[int]:
    """
    Suffix parsing rules:
      lorry_3t, lorry_10t  → strip trailing 't' → 3, 10
      trailer_20, trailer_40 → plain integer suffix → 20, 40
    """
    rows = conn.execute(text(
        "SELECT vehicle_type_id FROM vehicle_types WHERE is_active = true"
    )).fetchall()
    known = set()
    for (vt_id,) in rows:
        try:
            suffix = vt_id.split("_")[-1]
            if suffix.endswith("t"):
                known.add(int(suffix[:-1]))
            else:
                known.add(int(suffix))
        except (ValueError, IndexError):
            pass
    return known


# ---------------------------------------------------------------------------
# Step 1: Collect skipped cards from Datastore
# ---------------------------------------------------------------------------

def collect_skipped_cards(ds_client, areas_set: set[str], known_tonnage: set[int]) -> tuple[list[dict], set[int]]:
    logger.info("\n=== Step 1: Collecting skipped rate cards from Datastore ===")
    query = ds_client.query(kind=KIND_PRICING_TRANSPORT)
    entities = list(query.fetch())
    logger.info(f"  Fetched {len(entities)} PricingTransport entities")

    skipped_records: list[dict] = []
    skipped_pt_ids: set[int] = set()
    counts = {"trash": 0, "seeded": 0, "no_city_code": 0, "tonnage_not_found": 0, "area_not_found": 0, "ok": 0}

    for e in entities:
        pt_id = e.key.id_or_name
        if not pt_id:
            continue

        if e.get("trash", False):
            counts["trash"] += 1
            continue

        port_un_code = (e.get("port_un_code") or "").strip()
        city_code = (e.get("city_code") or "").strip()
        tonnage_raw = e.get("tonnage")
        include_dgf = bool(e.get("include_depot_gate_fee", False))

        if port_un_code in LEGACY_PORT_TERMINAL_MAP:
            port_un_code, _ = LEGACY_PORT_TERMINAL_MAP[port_un_code]

        if city_code in SEEDED_AREA_CODES:
            counts["seeded"] += 1
            continue

        country_code = city_code.split("-")[0] if city_code and "-" in city_code else ""

        try:
            tonnage_int = int(tonnage_raw) if tonnage_raw is not None else None
        except (ValueError, TypeError):
            tonnage_int = None

        if not city_code:
            skip_reason = "no_city_code"
            counts["no_city_code"] += 1
        elif tonnage_int is None or tonnage_int not in known_tonnage:
            skip_reason = "tonnage_not_found"
            counts["tonnage_not_found"] += 1
        elif city_code not in areas_set:
            skip_reason = "area_not_found"
            counts["area_not_found"] += 1
        else:
            counts["ok"] += 1
            continue

        skipped_records.append({
            "pt_id": pt_id,
            "port_un_code": port_un_code,
            "city_code": city_code,
            "country_code": country_code,
            "tonnage": tonnage_raw,
            "include_depot_gate_fee": include_dgf,
            "skip_reason": skip_reason,
            "action": "",
            "new_area_code": "",
            "new_area_name": "",
            "notes": "",
        })
        skipped_pt_ids.add(int(pt_id))

    logger.info(f"  Trashed (excluded):        {counts['trash']}")
    logger.info(f"  Seeded areas (excluded):   {counts['seeded']}")
    logger.info(f"  no_city_code:              {counts['no_city_code']}")
    logger.info(f"  tonnage_not_found:         {counts['tonnage_not_found']}")
    logger.info(f"  area_not_found:            {counts['area_not_found']}")
    logger.info(f"  Migrated OK:               {counts['ok']}")
    logger.info(f"  Total skipped for export:  {len(skipped_records)}")
    return skipped_records, skipped_pt_ids


# ---------------------------------------------------------------------------
# Step 2: Fetch City names from Datastore
# ---------------------------------------------------------------------------

def fetch_city_names(ds_client, city_codes: list[str]) -> dict[str, str]:
    """
    Fetch City entities from Datastore and return {city_code: name}.
    Falls back to the city_code itself if not found.
    """
    logger.info(f"\n=== Step 2: Fetching City names from Datastore ({len(city_codes)} codes) ===")
    keys = [ds_client.key(KIND_CITY, code) for code in city_codes]
    entities = _get_multi_chunked(ds_client, keys)

    city_map: dict[str, str] = {}
    for entity in entities:
        code = entity.key.name
        name = entity.get("name") or code
        city_map[code] = name

    found = len(city_map)
    missing = len(city_codes) - found
    logger.info(f"  Found: {found}, Not in Datastore: {missing}")
    if missing > 0:
        missing_codes = [c for c in city_codes if c not in city_map]
        for c in missing_codes:
            logger.info(f"    Missing: {c}")

    return city_map


# ---------------------------------------------------------------------------
# Step 3: Collect 2024+ rates for skipped cards
# ---------------------------------------------------------------------------

def collect_skipped_rates(ds_client, skipped_pt_ids: set[int], skipped_cards: list[dict]) -> list[dict]:
    logger.info("\n=== Step 3: Collecting 2024+ rates for skipped cards ===")

    card_context: dict[int, dict] = {
        int(c["pt_id"]): {
            "port_un_code": c["port_un_code"],
            "city_code": c["city_code"],
            "skip_reason": c["skip_reason"],
        }
        for c in skipped_cards
    }

    query = ds_client.query(kind=KIND_MONTHLY_RATE)
    query.add_filter(filter=PropertyFilter("kind", "=", PT_TRANSPORT))
    entities = list(query.fetch())
    logger.info(f"  Fetched {len(entities)} PT-TRANSPORT rate entities")

    rate_rows: list[dict] = []
    skipped_pre2024 = 0
    skipped_no_match = 0
    parse_errors = 0

    for e in entities:
        effective_from = _parse_month_year(e.get("month_year", ""))
        if not effective_from:
            parse_errors += 1
            continue
        if effective_from.year < 2024:
            skipped_pre2024 += 1
            continue

        try:
            pt_id = int(e.get("pt_id"))
        except (TypeError, ValueError):
            parse_errors += 1
            continue

        if pt_id not in skipped_pt_ids:
            skipped_no_match += 1
            continue

        ctx = card_context.get(pt_id, {})
        is_price = bool(e.get("is_price", False))
        price_data = e.get("price") or {}
        cost_data = e.get("cost") or {}
        charges_data = e.get("charges") or {}

        rate_rows.append({
            "pt_id": pt_id,
            "port_un_code": ctx.get("port_un_code", ""),
            "city_code": ctx.get("city_code", ""),
            "skip_reason": ctx.get("skip_reason", ""),
            "month_year": e.get("month_year", ""),
            "is_price": is_price,
            "supplier_id": "" if is_price else (e.get("supplier_id") or ""),
            "list_price": _safe_float(price_data.get("price")) if is_price else None,
            "min_list_price": _safe_float(price_data.get("min_price")) if is_price else None,
            "cost": _safe_float(cost_data.get("cost")) if not is_price else None,
            "min_cost": _safe_float(cost_data.get("min_cost")) if not is_price else None,
            "toll_fee": _safe_float(charges_data.get("toll_fee")),
            "side_loader_surcharge": _safe_float(charges_data.get("side_loader_surcharge")),
            "currency": (e.get("currency") or "MYR").strip(),
            "uom": (e.get("uom") or "SET").strip(),
            "roundup_qty": e.get("roundup_qty") or 0,
        })

    logger.info(f"  Skipped (pre-2024):         {skipped_pre2024}")
    logger.info(f"  Skipped (no matching card): {skipped_no_match}")
    logger.info(f"  Parse errors:               {parse_errors}")
    logger.info(f"  Rate rows for export:       {len(rate_rows)}")
    return rate_rows


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------

def _apply_header(ws, columns: list[tuple]):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def write_areas_excel(skipped_cards: list[dict], city_map: dict[str, str], output_dir: str, date_str: str) -> str:
    # Derive unique city_codes preserving first-seen order, sorted by country then code
    seen: dict[str, str] = {}
    for card in skipped_cards:
        cc = card["city_code"]
        if cc and cc not in seen:
            seen[cc] = card["country_code"]
    unique_areas = sorted(seen.items(), key=lambda x: (x[1], x[0]))

    filepath = os.path.join(output_dir, f"transport_skipped_areas_{date_str}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Unique Area Codes"

    columns = [("city_code", 22), ("country_code", 14), ("area_name", 36), ("notes", 30)]
    _apply_header(ws, columns)

    editable_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    for row_idx, (city_code, country_code) in enumerate(unique_areas, start=2):
        ws.cell(row=row_idx, column=1, value=city_code)
        ws.cell(row=row_idx, column=2, value=country_code)
        # Populate area_name from Datastore City kind; yellow if missing
        name = city_map.get(city_code, "")
        name_cell = ws.cell(row=row_idx, column=3, value=name)
        if not name:
            name_cell.fill = editable_fill
        notes_cell = ws.cell(row=row_idx, column=4, value="")
        notes_cell.fill = editable_fill

    ws_notes = wb.create_sheet("Instructions")
    ws_notes.column_dimensions["A"].width = 22
    ws_notes.column_dimensions["B"].width = 70
    for r_idx, (a, b) in enumerate([
        ("Column", "Purpose"),
        ("city_code", "Legacy area code — unique values across all skipped cards"),
        ("country_code", "Derived from city_code prefix — sorted by country for easier batching"),
        ("area_name", "Name from Datastore City kind. Yellow = not found in Datastore, fill manually."),
        ("notes", "Free text — decisions, corrections, ops notes"),
    ], start=1):
        ws_notes.cell(row=r_idx, column=1, value=a).font = Font(bold=True) if r_idx == 1 else Font()
        ws_notes.cell(row=r_idx, column=2, value=b)

    wb.save(filepath)
    logger.info(f"\n  Areas exported: {len(unique_areas)} unique city_codes → {filepath}")
    return filepath


def write_cards_excel(skipped_cards: list[dict], output_dir: str, date_str: str) -> str:
    filepath = os.path.join(output_dir, f"transport_skipped_cards_{date_str}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Skipped Rate Cards"

    columns = [
        ("pt_id", 12), ("port_un_code", 14), ("city_code", 18), ("country_code", 14),
        ("tonnage", 10), ("include_depot_gate_fee", 22), ("skip_reason", 20),
        ("action", 16), ("new_area_code", 18), ("new_area_name", 24), ("notes", 30),
    ]
    keys = ["pt_id", "port_un_code", "city_code", "country_code", "tonnage",
            "include_depot_gate_fee", "skip_reason", "action", "new_area_code", "new_area_name", "notes"]
    _apply_header(ws, columns)

    editable_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    editable_keys = {"action", "new_area_code", "new_area_name", "notes"}
    editable_cols = {i + 1 for i, k in enumerate(keys) if k in editable_keys}

    for row_idx, record in enumerate(skipped_cards, start=2):
        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))
            if col_idx in editable_cols:
                cell.fill = editable_fill

    ws_notes = wb.create_sheet("Instructions")
    ws_notes.column_dimensions["A"].width = 28
    ws_notes.column_dimensions["B"].width = 60
    for r_idx, (a, b) in enumerate([
        ("Column", "Purpose"),
        ("pt_id", "Datastore entity ID — join key to rates file"),
        ("skip_reason", "area_not_found / tonnage_not_found / no_city_code"),
        ("action", "Fill in: seed_area / map_existing / drop / defer"),
        ("new_area_code", "Proposed standardised area code"),
        ("new_area_name", "Human-readable area name for new seed"),
        ("", ""),
        ("Action values", ""),
        ("seed_area", "Create a new area row with new_area_code + new_area_name"),
        ("map_existing", "Map city_code to an existing area — put target code in new_area_code"),
        ("drop", "Intentionally exclude — data no longer relevant"),
        ("defer", "No decision yet — revisit later"),
    ], start=1):
        ws_notes.cell(row=r_idx, column=1, value=a).font = Font(bold=True) if r_idx == 1 else Font()
        ws_notes.cell(row=r_idx, column=2, value=b)

    wb.save(filepath)
    logger.info(f"  Cards exported: {len(skipped_cards)} rows → {filepath}")
    return filepath


def write_rates_excel(skipped_rates: list[dict], output_dir: str, date_str: str) -> str:
    filepath = os.path.join(output_dir, f"transport_skipped_rates_{date_str}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Skipped Rates"

    columns = [
        ("pt_id", 16), ("port_un_code", 14), ("city_code", 20), ("skip_reason", 20),
        ("month_year", 12), ("is_price", 10), ("supplier_id", 20),
        ("list_price", 14), ("min_list_price", 16), ("cost", 14), ("min_cost", 14),
        ("toll_fee", 12), ("side_loader_surcharge", 22),
        ("currency", 10), ("uom", 10), ("roundup_qty", 14),
    ]
    keys = ["pt_id", "port_un_code", "city_code", "skip_reason", "month_year", "is_price",
            "supplier_id", "list_price", "min_list_price", "cost", "min_cost",
            "toll_fee", "side_loader_surcharge", "currency", "uom", "roundup_qty"]
    _apply_header(ws, columns)

    for row_idx, record in enumerate(skipped_rates, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

    ws_notes = wb.create_sheet("Instructions")
    ws_notes.column_dimensions["A"].width = 80
    for r_idx, line in enumerate([
        "2024+ rate rows for skipped cards. Use pt_id to join with the cards file.",
        "is_price=True: list_price/min_list_price populated, cost/min_cost blank.",
        "is_price=False: cost/min_cost populated, list_price/min_list_price blank.",
        "supplier_id is blank for is_price=True rows.",
    ], start=1):
        ws_notes.cell(row=r_idx, column=1, value=line)

    wb.save(filepath)
    logger.info(f"  Rates exported: {len(skipped_rates)} rows → {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    key_file = os.path.join(os.path.dirname(__file__), '..', 'cloud-accele-freight-key.json')
    if os.path.exists(key_file):
        os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = os.path.abspath(key_file)

    engine = get_engine()
    with engine.connect() as conn:
        areas_set = load_areas_set(conn)
        known_tonnage = load_known_tonnage_set(conn)
    logger.info(f"  Loaded {len(areas_set)} known area codes from DB")
    logger.info(f"  Loaded known tonnage values: {sorted(known_tonnage)}")

    ds_client = datastore.Client(project=PROJECT_ID)

    skipped_cards, skipped_pt_ids = collect_skipped_cards(ds_client, areas_set, known_tonnage)

    # Fetch City names for all unique skipped city_codes
    unique_city_codes = list({c["city_code"] for c in skipped_cards if c["city_code"]})
    city_map = fetch_city_names(ds_client, unique_city_codes)

    skipped_rates = collect_skipped_rates(ds_client, skipped_pt_ids, skipped_cards)

    output_dir = os.path.join(os.path.dirname(__file__), "output")
    os.makedirs(output_dir, exist_ok=True)
    date_str = date_cls.today().strftime('%Y%m%d')

    areas_path = write_areas_excel(skipped_cards, city_map, output_dir, date_str)
    cards_path = write_cards_excel(skipped_cards, output_dir, date_str)
    rates_path = write_rates_excel(skipped_rates, output_dir, date_str)

    logger.info(f"\n=== Export complete ===")
    logger.info(f"  Areas: {areas_path}")
    logger.info(f"  Cards: {cards_path} ({len(skipped_cards)} rows)")
    logger.info(f"  Rates: {rates_path} ({len(skipped_rates)} rows)")


if __name__ == "__main__":
    main()
