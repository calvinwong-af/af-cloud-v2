"""
scripts/export_faf_to_excel.py

Export FAF data from PT-HAULAGE rate entities (Datastore) to Excel.
Excludes all Malaysian ports (MY*, MYPKG, MYPKG_N).

Columns:
  - supplier_id
  - port_un_code (resolved via pt_id → PricingHaulage card lookup)
  - month_year
  - area_id (legacy city_code from PricingHaulage card)
  - equipment_size  (container_size from PricingHaulage card)
  - equipment_type  (vehicle_type or equivalent field from rate entity / card)
  - faf_percent
  - faf_value
  - has_faf_percent (boolean)
  - has_faf_value (boolean)

Run from af-server root with venv active:
    .venv\\Scripts\\python scripts\\export_faf_to_excel.py
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

key_file = os.path.join(os.path.dirname(__file__), '..', 'cloud-accele-freight-key.json')
if os.path.exists(key_file):
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = os.path.abspath(key_file)

from google.cloud import datastore
from google.cloud.datastore.query import PropertyFilter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl not installed. Run: .venv\\Scripts\\pip install openpyxl")
    sys.exit(1)

PROJECT_ID = 'cloud-accele-freight'

MY_PORT_PREFIXES = ("MY",)  # catches MY, MYPKG, MYPKG_N, MYPEN, etc.

def _safe_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None

print("Connecting to Datastore...")
ds = datastore.Client(project=PROJECT_ID)

# Step 1: Build pt_id → card info map from PricingHaulage cards
print("Loading PricingHaulage cards...")
cards = list(ds.query(kind="PricingHaulage").fetch())
print(f"  {len(cards)} cards loaded")

pt_id_to_card: dict[int, dict] = {}
for e in cards:
    pt_id = e.key.id_or_name
    port = (e.get("port_un_code") or "").strip()
    if pt_id and port:
        try:
            pt_id_to_card[int(pt_id)] = {
                "port_un_code":   port,
                "area_id":        (e.get("city_code") or "").strip(),  # legacy area identifier
                "equipment_size": (e.get("container_size") or e.get("equipment_size") or "").strip(),
                "equipment_type": (e.get("vehicle_type") or e.get("equipment_type") or e.get("transport_type") or "").strip(),
            }
        except (TypeError, ValueError):
            pass

# Step 2: Fetch all PT-HAULAGE rate entities
print("Loading PT-HAULAGE rate entities...")
rate_query = ds.query(kind="PTMonthlyRateHaulageTransport")
rate_query.add_filter(filter=PropertyFilter("kind", "=", "PT-HAULAGE"))
entities = list(rate_query.fetch())
print(f"  {len(entities)} rate entities loaded")

# Step 3: Extract FAF rows, excluding MY ports
rows = []
skipped_my = 0
skipped_no_faf = 0
skipped_no_port = 0

for e in entities:
    faf_data = e.get("faf") or {}
    if not isinstance(faf_data, dict):
        skipped_no_faf += 1
        continue

    faf_percent = _safe_float(faf_data.get("faf_percent") or faf_data.get("percent"))
    faf_value   = _safe_float(faf_data.get("faf_value")   or faf_data.get("value"))

    # Skip if no meaningful FAF data
    if faf_percent is None and faf_value is None:
        skipped_no_faf += 1
        continue

    pt_id_raw = e.get("pt_id")
    try:
        pt_id = int(pt_id_raw)
    except (TypeError, ValueError):
        skipped_no_port += 1
        continue

    card_info = pt_id_to_card.get(pt_id)
    if not card_info:
        skipped_no_port += 1
        continue

    port = card_info["port_un_code"]

    # Exclude Malaysian ports
    if port.upper().startswith(MY_PORT_PREFIXES):
        skipped_my += 1
        continue

    is_price = bool(e.get("is_price", False))
    supplier_id = None if is_price else (e.get("supplier_id") or None)

    # Equipment size/type: prefer fields on the rate entity itself, fall back to card
    equipment_size = (
        (e.get("container_size") or e.get("equipment_size") or "").strip()
        or card_info["equipment_size"]
    )
    equipment_type = (
        (e.get("vehicle_type") or e.get("equipment_type") or e.get("transport_type") or "").strip()
        or card_info["equipment_type"]
    )

    rows.append({
        "supplier_id":      supplier_id or "(list price)",
        "port_un_code":     port,
        "month_year":       e.get("month_year") or "",
        "area_id":          card_info["area_id"],
        "equipment_size":   equipment_size,
        "equipment_type":   equipment_type,
        "faf_percent":      faf_percent,
        "faf_value":        faf_value,
        "has_faf_percent":  faf_percent is not None,
        "has_faf_value":    faf_value is not None,
    })

print(f"\nSummary:")
print(f"  Rows with non-zero FAF (non-MY): {len(rows)}")
print(f"  Skipped (Malaysian port):        {skipped_my}")
print(f"  Skipped (no FAF data):           {skipped_no_faf}")
print(f"  Skipped (no port resolved):      {skipped_no_port}")

if not rows:
    print("\nNo non-Malaysian FAF data found. Nothing to export.")
    sys.exit(0)

# Sort by port, then supplier, then month_year
rows.sort(key=lambda r: (r["port_un_code"], r["supplier_id"], r["month_year"]))

# Step 4: Write to Excel
output_dir = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(output_dir, exist_ok=True)
out_path = os.path.join(output_dir, "faf_export_non_my.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FAF Non-MY"

headers = [
    "supplier_id",
    "port_un_code",
    "month_year",
    "area_id",
    "equipment_size",
    "equipment_type",
    "faf_percent",
    "faf_value",
    "has_faf_percent",
    "has_faf_value",
]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row_idx, row in enumerate(rows, 2):
    ws.cell(row=row_idx, column=1,  value=row["supplier_id"])
    ws.cell(row=row_idx, column=2,  value=row["port_un_code"])
    ws.cell(row=row_idx, column=3,  value=row["month_year"])
    ws.cell(row=row_idx, column=4,  value=row["area_id"])
    ws.cell(row=row_idx, column=5,  value=row["equipment_size"])
    ws.cell(row=row_idx, column=6,  value=row["equipment_type"])
    ws.cell(row=row_idx, column=7,  value=row["faf_percent"])
    ws.cell(row=row_idx, column=8,  value=row["faf_value"])
    ws.cell(row=row_idx, column=9,  value=row["has_faf_percent"])
    ws.cell(row=row_idx, column=10, value=row["has_faf_value"])

col_widths = [22, 15, 12, 22, 16, 16, 14, 12, 16, 14]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(1, col).column_letter].width = width

ws.freeze_panes = "A2"

wb.save(out_path)
print(f"\nExported to: {out_path}")
print(f"Rows written: {len(rows)}")
